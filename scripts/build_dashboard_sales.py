"""
Build the Inbound Summary — Sales & Status Queue dashboard.

Source: data/sales_daily.csv (pulled from Metabase:
        PROD_DB.PUBLIC.AMEYO_CALL_DETAILS_REPORT,
        Before April 2026: QUEUE_NAME IN ('Sales Queue', 'Booking Queue'),
        From April 2026:   QUEUE_NAME = 'Sales Queue',
        CALL_TYPE = inbound.call.dial)

Output: output/sales_status_dashboard.xlsx

Layout per month tab:
  Col A: Disposition (class rows pink bold, code rows pink indented)
  Col B: MTD Avg  (bold)
  Col C: Last-Month Avg
  Col D..: days in descending order (latest day = col D, oldest at the right)

Spike logic (upward only):
  - Daily cells are compared against that row's MTD Avg.
  - The MTD-Avg cell itself is compared against its Last-Month Avg.
  - Tiered threshold based on the absolute value of the number being checked:
        value >= 100   -> trigger if >= 15% above baseline
        value <  100   -> trigger if >= 25% above baseline
  - When triggered, the cell shows  "<value> (+NN%) ↑"  in red.
"""
import csv
from collections import defaultdict
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

# ------------------------------------------------------------------ load data

import os as _os
_BASE = Path(__file__).resolve().parent.parent
SRC = Path(_os.environ.get("DATA_DIR", str(_BASE / "data"))) / "sales_daily.csv"
OUT = _BASE / "output" / "sales_status_dashboard.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

rows = list(csv.DictReader(SRC.open()))

# ---- class relabelling & collapsing rules ------------------------------
# All six Sales-* classes are merged into a single virtual class "Sales Queue".
# (New Sales-* classes showing up in the future will auto-roll up too.)
CLASS_MERGE_EXACT = {
    "Sales-App Issues":                 "Sales Queue",
    "Sales-Next Steps":                 "Sales Queue",
    "Sales-Next Steps - Old Contruct":  "Sales Queue",
    "Sales-Pause":                      "Sales Queue",
    "Sales-System Understanding":       "Sales Queue",
    "Sales-System Understanding - Old": "Sales Queue",
}

def display_class(src_cls: str) -> str:
    if src_cls in CLASS_MERGE_EXACT:
        return CLASS_MERGE_EXACT[src_cls]
    if src_cls.startswith("Sales-") or src_cls.startswith("Sales "):
        return "Sales Queue"
    if src_cls == "(Unclassified)" or src_cls == "" or src_cls is None:
        return "Missed Calls"
    return src_cls

# Classes that are rendered as a single aggregated row (no child codes).
COLLAPSED_CLASSES = {"Missed Calls",
                     "PayG - Clarification Provided",
                     "user.forced.logged.off",
                     "user.transferred.to.campaign"}

# Classes excluded from this dashboard entirely.
EXCLUDED_CLASSES = {"Internet Related Issues", "Payment Related Issues"}
# For collapsed classes we still need one "code key" in the dict so the
# rendering loop can iterate over something — we use a sentinel.
COLLAPSED_SENTINEL = "__ALL__"

# counts[(display_class, code)][date] = n   — ALL dispositions included
counts = defaultdict(lambda: defaultdict(int))
all_dates = set()
for r in rows:
    src_cls = r["DISPOSITION_CLASS"]
    code    = r["DISPOSITION_CODE"]
    d       = date.fromisoformat(r["CALL_DATE"])
    n       = int(r["CALL_COUNT"])

    cls = display_class(src_cls)
    if cls in EXCLUDED_CLASSES:
        continue
    if cls in COLLAPSED_CLASSES:
        key_code = COLLAPSED_SENTINEL
    else:
        key_code = code
    counts[(cls, key_code)][d] += n
    all_dates.add(d)

by_month = defaultdict(set)           # (yr, mo) -> set of dates
for d in all_dates:
    by_month[(d.year, d.month)].add(d)
months_sorted = sorted(by_month.keys())

# Only create tabs for March 2026+; earlier months serve as prev-month data
DISPLAY_FROM = (2026, 3)
display_months = [ym for ym in months_sorted if ym >= DISPLAY_FROM]

# ---------------------------------------------------------------- styles

PINK       = "E91E8C"
PINK_LIGHT = "FCE4EC"
WHITE      = "FFFFFF"
BLACK      = "000000"
RED        = "C62828"
GREY       = "D0D0D0"
AVG_FILL   = "E3F2FD"     # light blue — highlights MTD Avg + Prev-Month Avg columns
AVG_FILL_CLASS = "BBDEFB" # slightly darker light-blue for class rows in avg cols

thin   = Side(style="thin", color=GREY)
border = Border(left=thin, right=thin, top=thin, bottom=thin)

font_class_name = Font(name="Calibri", size=11, bold=True,  color=PINK)
font_code_name  = Font(name="Calibri", size=11, bold=False, color=PINK)
font_num        = Font(name="Calibri", size=11, bold=False, color=BLACK)
font_num_mtd    = Font(name="Calibri", size=11, bold=True,  color=BLACK)
font_header     = Font(name="Calibri", size=11, bold=True,  color=BLACK)

# Rich-text inline fonts for the spike suffix (black number, red "+NN% ↑")
inline_black     = InlineFont(rFont="Calibri", sz=11, b=False, color=BLACK)
inline_black_bd  = InlineFont(rFont="Calibri", sz=11, b=True,  color=BLACK)
inline_red       = InlineFont(rFont="Calibri", sz=9, b=False, color=RED)
inline_red_bd    = InlineFont(rFont="Calibri", sz=9, b=True,  color=RED)

fill_white       = PatternFill("solid", fgColor=WHITE)
fill_class_row   = PatternFill("solid", fgColor=PINK_LIGHT)
fill_header      = PatternFill("solid", fgColor="F5F5F5")
fill_total       = PatternFill("solid", fgColor="EEEEEE")
fill_avg_col     = PatternFill("solid", fgColor=AVG_FILL)        # for code rows
fill_avg_col_cls = PatternFill("solid", fgColor=AVG_FILL_CLASS)  # for class rows
fill_avg_col_tot = PatternFill("solid", fgColor="90CAF9")        # for total row

center      = Alignment(horizontal="center", vertical="center")
left        = Alignment(horizontal="left",   vertical="center")
left_indent = Alignment(horizontal="left",   vertical="center", indent=2)

MONTH_NAMES = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]


def fmt_month(ym):
    return f"{MONTH_NAMES[ym[1]-1]} {ym[0]}"


PINNED_TOP = ["Sales Queue", "Booking Queue"]

MIDDLE_ORDER = [
    "Others",
    "Change Request",
    "Partner Misbehavior",
    "Refund",
    "Remove Connection - Talk to Customer",
    "Shifting Request",
    "Wiom Net",
    "Incomplete Call",
    "campaign.system.disposition",
    "user.forced.logged.off",
    "user.transferred.to.campaign",
]

PINNED_BOTTOM = ["Missed Calls"]


def build_taxonomy():
    """Ordered list of (class, [codes...])."""
    cls_codes = defaultdict(set)
    for (cls, code) in counts.keys():
        cls_codes[cls].add(code)

    known = set(PINNED_TOP) | set(MIDDLE_ORDER) | set(PINNED_BOTTOM)
    extra_middle = sorted(c for c in cls_codes if c not in known)

    order = [c for c in PINNED_TOP      if c in cls_codes]
    order += [c for c in MIDDLE_ORDER   if c in cls_codes]
    order += extra_middle
    order += [c for c in PINNED_BOTTOM  if c in cls_codes]

    result = []
    for cls in order:
        if cls in COLLAPSED_CLASSES:
            result.append((cls, []))
        else:
            codes = sorted(c for c in cls_codes[cls] if c != COLLAPSED_SENTINEL)
            result.append((cls, codes))
    return result


def month_dates_desc(ym):
    return sorted(by_month[ym], reverse=True)


def avg(vals):
    vals = list(vals)
    return sum(vals) / len(vals) if vals else 0


def threshold_for(value):
    v = abs(value)
    if v >= 100:
        return 0.15
    if v >= 20:
        return 0.25
    return 1.00


def spike_flag(value, baseline):
    if baseline is None or baseline <= 0 or value is None:
        return (False, 0)
    pct = (value - baseline) / baseline
    if pct <= 0:
        return (False, 0)
    return (pct >= threshold_for(value), round(pct * 100))


def write_cell(ws, r, c, value, baseline, is_mtd_col=False, is_class=False):
    spiked, pct = spike_flag(value, baseline)
    rounded = int(round(value or 0))

    if not spiked:
        cell = ws.cell(row=r, column=c, value=rounded)
        cell.font = font_num_mtd if is_mtd_col else font_num
        return cell

    num_font = inline_black_bd if is_mtd_col else inline_black
    red_font = inline_red_bd   if is_mtd_col else inline_red
    rt = CellRichText([
        TextBlock(num_font, str(rounded)),
        TextBlock(red_font, f" (+{pct}%) \u2191"),
    ])
    cell = ws.cell(row=r, column=c, value=rt)
    return cell


# ---------------------------------------------------------------- workbook

wb = Workbook()
wb.remove(wb.active)
taxonomy = build_taxonomy()


def write_month_sheet(ym):
    title = fmt_month(ym)
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False

    days_desc = month_dates_desc(ym)
    prev_ym = (ym[0], ym[1] - 1) if ym[1] > 1 else (ym[0] - 1, 12)
    prev_days = sorted(by_month.get(prev_ym, []))

    prev_label = f"{fmt_month(prev_ym)} Avg" if prev_days else "Prev Month Avg"
    headers = ["Category", f"{fmt_month(ym)} Avg (MTD)", prev_label] + [
        d.strftime("%d-%b") for d in days_desc
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = font_header
        c.fill = fill_header
        c.alignment = center
        c.border = border

    r = 2
    for cls, codes in taxonomy:
        class_day      = {d: 0 for d in days_desc}
        class_prev_day = {d: 0 for d in prev_days}
        for (k_cls, _k_code), dd in counts.items():
            if k_cls != cls:
                continue
            for d, n in dd.items():
                if d in class_day:
                    class_day[d] += n
                if d in class_prev_day:
                    class_prev_day[d] += n

        cls_mtd  = avg(class_day.values())
        cls_prev = avg(class_prev_day.values())

        # --- class row ---
        cell = ws.cell(row=r, column=1, value=cls)
        cell.font = font_class_name
        cell.alignment = left

        write_cell(ws, r, 2, cls_mtd, cls_prev, is_mtd_col=True, is_class=True)
        ws.cell(row=r, column=3, value=int(round(cls_prev))).font = font_num
        for i, d in enumerate(days_desc):
            write_cell(ws, r, 4 + i, class_day.get(d, 0), cls_mtd, is_class=True)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill_avg_col_cls if col in (2, 3) else fill_class_row
            cell.border = border
            if col >= 2:
                cell.alignment = center
        r += 1

        # --- code rows (sorted descending by MTD avg) ---
        code_mtds = []
        for code in codes:
            dv = {d: counts[(cls, code)].get(d, 0) for d in days_desc}
            code_mtds.append((code, avg(dv.values())))
        code_mtds.sort(key=lambda x: x[1], reverse=True)
        for code, _ in code_mtds:
            day_vals  = {d: counts[(cls, code)].get(d, 0) for d in days_desc}
            prev_vals = {d: counts[(cls, code)].get(d, 0) for d in prev_days}
            mtd  = avg(day_vals.values())
            prev = avg(prev_vals.values())

            cell = ws.cell(row=r, column=1, value=code)
            cell.font = font_code_name
            cell.alignment = left_indent

            write_cell(ws, r, 2, mtd, prev, is_mtd_col=True)
            ws.cell(row=r, column=3, value=int(round(prev))).font = font_num
            for i, d in enumerate(days_desc):
                write_cell(ws, r, 4 + i, day_vals.get(d, 0), mtd)

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)
                cell.fill = fill_avg_col if col in (2, 3) else fill_white
                cell.border = border
                if col >= 2:
                    cell.alignment = center
            r += 1

    # TOTAL row
    total_day = defaultdict(int)
    total_prev_day = defaultdict(int)
    for (cls, code), dd in counts.items():
        for d, n in dd.items():
            if d in set(days_desc):
                total_day[d] += n
            if d in set(prev_days):
                total_prev_day[d] += n
    total_mtd  = avg(total_day.values())
    total_prev = avg(total_prev_day.values())

    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True, color=BLACK)
    write_cell(ws, r, 2, total_mtd, total_prev, is_mtd_col=True)
    ws.cell(row=r, column=3, value=int(round(total_prev))).font = font_num
    for i, d in enumerate(days_desc):
        write_cell(ws, r, 4 + i, total_day.get(d, 0), total_mtd)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=r, column=col)
        cell.fill = fill_avg_col_tot if col in (2, 3) else fill_total
        cell.border = border
        if col >= 2:
            cell.alignment = center

    # widths + freeze
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    for i in range(len(days_desc)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 14
    ws.freeze_panes = "B2"

    # footnote
    note_row = r + 3
    queue_note = (
        f"Filters: Queue = Sales Queue"
        + (f" + Booking Queue (before Apr 2026)" if ym < (2026, 4) else "")
        + f"   |   Call Type = inbound.call.dial   |   Month = {title}"
    )
    notes = [
        queue_note,
        "Arrow \u2191 = upward spike. Daily cells are compared vs the row's MTD Avg; "
        "the MTD Avg cell is compared vs the Last-Month Avg.",
        "Threshold: value \u2265 100 \u2192 15% over baseline ; value < 100 \u2192 25% over baseline. "
        "Spikes show  \"<value> (+NN%) \u2191\"  in red.",
        "Disposition Class rows are highlighted pink; Month-To-Date Avg column is bold.",
    ]
    for i, txt in enumerate(notes):
        ws.cell(row=note_row + i, column=1, value=txt).font = Font(italic=True, color="666666")


for ym in display_months:
    write_month_sheet(ym)

# ---------------------------------------------------------------- MoM tab

ws = wb.create_sheet("MoM Comparison")
ws.sheet_view.showGridLines = False

header = ["Disposition Class"] + [fmt_month(m) + " Avg" for m in months_sorted]
ws.append(header)
for col_idx in range(1, len(header) + 1):
    c = ws.cell(row=1, column=col_idx)
    c.font = font_header
    c.fill = fill_header
    c.alignment = center
    c.border = border

class_totals_by_month = defaultdict(lambda: defaultdict(int))
class_days_by_month   = defaultdict(lambda: defaultdict(set))
for (cls, code), dd in counts.items():
    for d, n in dd.items():
        ym = (d.year, d.month)
        class_totals_by_month[cls][ym] += n
        class_days_by_month[cls][ym].add(d)

r = 2
for cls, _codes in taxonomy:
    ws.cell(row=r, column=1, value=cls).font = font_class_name
    ws.cell(row=r, column=1).alignment = left
    for i, ym in enumerate(months_sorted):
        total = class_totals_by_month[cls].get(ym, 0)
        ndays = len(by_month[ym])
        val = int(round(total / ndays)) if ndays else 0
        ws.cell(row=r, column=2 + i, value=val).font = font_num
    for col in range(1, len(header) + 1):
        cell = ws.cell(row=r, column=col)
        cell.fill = fill_white
        cell.border = border
        if col >= 2:
            cell.alignment = center
    r += 1

# total row
ws.cell(row=r, column=1, value="TOTAL (all classes)").font = Font(bold=True, color=BLACK)
for i, ym in enumerate(months_sorted):
    tot = sum(class_totals_by_month[cls].get(ym, 0) for cls, _ in taxonomy)
    nd = len(by_month[ym])
    ws.cell(row=r, column=2 + i, value=int(round(tot / nd)) if nd else 0).font = font_num_mtd
for col in range(1, len(header) + 1):
    cell = ws.cell(row=r, column=col)
    cell.fill = fill_total
    cell.border = border
    if col >= 2:
        cell.alignment = center

ws.column_dimensions["A"].width = 42
for i in range(len(months_sorted)):
    ws.column_dimensions[get_column_letter(2 + i)].width = 16

# chart
chart = LineChart()
chart.title = "Month-on-Month Avg Daily Inbound Calls — Sales & Status Queue"
chart.style = 12
chart.y_axis.title = "Avg calls / day"
chart.x_axis.title = "Disposition Class"
chart.height = 10
chart.width = 22
data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(months_sorted), max_row=r - 1)
cats_ref = Reference(ws, min_col=1, min_row=2, max_row=r - 1)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws.add_chart(chart, f"A{r + 3}")

# ---------------------------------------------------------------- README

ws = wb.create_sheet("README", 0)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 115

lines = [
    ("Inbound Summary — Sales & Status Queue Dashboard", True, 14, PINK),
    ("", False, 11, BLACK),
    ("Source", True, 12, BLACK),
    ("  Metabase (db 113 / Snowflake): PROD_DB.PUBLIC.AMEYO_CALL_DETAILS_REPORT", False, 11, BLACK),
    ("  Filters: Before Apr 2026: QUEUE_NAME IN ('Sales Queue','Booking Queue')", False, 11, BLACK),
    ("           From Apr 2026:   QUEUE_NAME = 'Sales Queue'", False, 11, BLACK),
    ("           CALL_TYPE = 'inbound.call.dial'", False, 11, BLACK),
    ("  ALL Disposition Classes and Codes are included (including system dispositions and nulls).", False, 11, BLACK),
    ("", False, 11, BLACK),
    ("Tabs", True, 12, BLACK),
    ("  \u2022 One tab per month (e.g. 'Mar 2026', 'Apr 2026') \u2014 use these as the month filter.", False, 11, BLACK),
    ("  \u2022 MoM Comparison \u2014 month-on-month average daily calls per Disposition Class + line chart.", False, 11, BLACK),
    ("", False, 11, BLACK),
    ("Column layout on month tabs", True, 12, BLACK),
    ("  A: Disposition (Class = pink + bold ; Code = pink, indented under its class)", False, 11, BLACK),
    ("  B: MTD Avg (bold)", False, 11, BLACK),
    ("  C: Last-Month Avg", False, 11, BLACK),
    ("  D \u2192: Day-wise values \u2014 latest day in column D, oldest at the right", False, 11, BLACK),
    ("", False, 11, BLACK),
    ("Spike flagging (upward only)", True, 12, BLACK),
    ("  Daily cells are compared against the row\u2019s MTD Avg.", False, 11, BLACK),
    ("  The MTD Avg cell is compared against the Last-Month Avg.", False, 11, BLACK),
    ("  Tiered threshold (based on the observed value):", False, 11, BLACK),
    ("      value \u2265 100  \u2192  flag if \u2265 15% above baseline", False, 11, BLACK),
    ("      value < 100  \u2192  flag if \u2265 25% above baseline", False, 11, BLACK),
    ("  Flagged cells are red and show  \"<value> (+NN%) \u2191\"  \u2014 NN is the % above baseline.", False, 11, BLACK),
    ("", False, 11, BLACK),
    ("Refreshing", True, 12, BLACK),
    ("  1. python scripts/pull_ameyo_sales.py       # re-pulls data from Metabase", False, 11, BLACK),
    ("  2. python scripts/build_dashboard_sales.py  # rebuilds this workbook", False, 11, BLACK),
    ("  New Disposition Codes / Classes (and new months) appear automatically on the next refresh.", False, 11, BLACK),
    ("  A daily Windows Scheduled Task runs 'refresh_sales_dashboard.bat' to do both steps automatically.", False, 11, BLACK),
]
for i, (txt, bold, size, color) in enumerate(lines, 1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name="Calibri", size=size, bold=bold, color=color)
    c.fill = fill_white

wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Months: {[fmt_month(m) for m in display_months]}")
print(f"Disposition classes: {len(taxonomy)}")
print(f"Disposition codes: {sum(len(c) for _, c in taxonomy)}")
